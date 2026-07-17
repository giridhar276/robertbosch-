# Topic: Third-party Libraries
# Example: Read Excel report

import os

try:
    from openpyxl import load_workbook
except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    raise SystemExit

file_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), "outputs", "employee_report.xlsx")

if not os.path.exists(file_path):
    print("Run 01_openpyxl_create_excel_report.py first")
else:
    # load_workbook() opens an existing Excel workbook.
    workbook = load_workbook(file_path)
    sheet = workbook["Employees"]

    for row in sheet.iter_rows(values_only=True):
        print(row)
